import os
from openpyxl import load_workbook

# 1. 获取脚本目录，即游戏根目录
root_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(root_dir, "name.xlsx")

# 2. 读取 name.xlsx
if not os.path.exists(file_path):
    print(f"文件 {file_path} 不存在！")
    exit()

wb = load_workbook(file_path)
ws = wb.active  # 默认操作第一个sheet

# 获取第一列数据及行号
rows = list(ws.iter_rows(min_row=1, max_col=1, values_only=True))

# 用于标记需要删除的行
delete_rows = set()

print("原始数据：")
for idx, row in enumerate(rows, start=1):
    print(f"{idx}: {row[0]}")

# 3. 删除首字母为数字的行
for idx, row in enumerate(rows, start=1):
    if row[0] and str(row[0])[0].isdigit():
        delete_rows.add(idx)
        print(f"删除首字母为数字的行 {idx}: {row[0]}")

# 4. 删除只有一个字母的行
for idx, row in enumerate(rows, start=1):
    if idx in delete_rows:
        continue
    if row[0] and len(str(row[0]).strip()) == 1 and str(row[0]).isalpha():
        delete_rows.add(idx)
        print(f"删除只有一个字母的行 {idx}: {row[0]}")

# 5. 删除重复行，只保留唯一行
seen = set()
for idx, row in enumerate(rows, start=1):
    if idx in delete_rows:
        continue
    val = str(row[0]).strip() if row[0] else ""
    if val in seen:
        delete_rows.add(idx)
        print(f"删除重复行 {idx}: {row[0]}")
    else:
        seen.add(val)

# 删除行（必须从下往上删除，否则行号会变化）
for row_idx in sorted(delete_rows, reverse=True):
    ws.delete_rows(row_idx)

# 保存文件
wb.save(file_path)
print("处理完成，已保存。")

# 保持窗口打开，等待用户手动关闭
input("按回车键退出...")

import openpyxl
import re
import os

# 最终输出文件名
OUTPUT_FILENAME_POSITIVE = "result正.xlsx"
OUTPUT_FILENAME_NEGATIVE = "result反.xlsx"
INPUT_FILENAME = "result.txt"

def write_data_to_excel(filename, data_list, column_order, title="占位符数据"):
    """
    将解析后的数据写入指定的 Excel 文件，并允许自定义列的顺序。
    """
    try:
        print(f"正在写入文件: {filename} ...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title

        # 设置列宽
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 40

        # 写入数据
        for row_index, (label, placeholder) in enumerate(data_list, 1):
            sheet.cell(row=row_index, column=1, value=(label, placeholder)[column_order[0]])
            sheet.cell(row=row_index, column=2, value=(label, placeholder)[column_order[1]])

        # 保存文件
        workbook.save(filename)
        print(f"✅ 写入完成: {filename}")
        return True
    except Exception as e:
        print(f"❌ 写入文件 {filename} 失败: {e}")
        return False


def convert_txt_to_excel_dual(input_filename=INPUT_FILENAME):
    """
    读取 result.txt，解析内容，并生成 result正.xlsx 和 result反.xlsx。
    """
    print("="*50)
    print("开始 Excel 转换流程...")
    print("="*50)

    # 1. 读取文件
    print(f"\n步骤1: 读取文件: {input_filename}")
    try:
        with open(input_filename, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        print(f"✅ 成功读取 {len(lines)} 行内容")
    except FileNotFoundError:
        print(f"❌ 错误: 找不到文件 {input_filename}。")
        input("按回车键退出...")
        return
    except Exception as e:
        print(f"❌ 读取文件时发生错误: {e}")
        input("按回车键退出...")
        return

    # 2. 解析数据
    print("\n步骤2: 解析内容为 [内容] 和 Emoji 对")
    parsed_data = []
    pattern = re.compile(r'^\s*(.+?)\s+([^\s]+)\s*$', re.UNICODE)

    for line_number, line in enumerate(lines, 1):
        line = line.strip()
        if not line or line.startswith('#'):
            continue

        match = pattern.match(line)
        if match:
            content = match.group(1).strip()
            emoji = match.group(2).strip()
            parsed_data.append((content, emoji))
        else:
            print(f"⚠️ 警告: 第 {line_number} 行未匹配成功，跳过: {line}")

    if not parsed_data:
        print("❌ 未解析到任何有效数据，程序终止。")
        input("按回车键退出...")
        return

    print(f"✅ 数据解析完成，共 {len(parsed_data)} 条有效记录。")

    # 3. 生成 Excel 文件
    print("\n步骤3: 生成 Excel 文件")

    # 3.1 result正.xlsx (正向: 左列 [内容], 右列 Emoji)
    print("\n生成 result正.xlsx (正向)")
    success_positive = write_data_to_excel(
        OUTPUT_FILENAME_POSITIVE,
        parsed_data,
        column_order=(0, 1),
        title="正向对照"
    )

    # 3.2 result反.xlsx (反向: 左列 Emoji, 右列 [内容])
    print("\n生成 result反.xlsx (反向)")
    success_negative = write_data_to_excel(
        OUTPUT_FILENAME_NEGATIVE,
        parsed_data,
        column_order=(1, 0),
        title="反向对照"
    )

    print("\n" + "="*50)
    if success_positive and success_negative:
        print(f"✅ Excel 文件生成完成！")
    else:
        print(f"❌ Excel 文件生成过程中出现问题，请检查日志。")
    print(f"正向文件: {os.path.abspath(OUTPUT_FILENAME_POSITIVE)}")
    print(f"反向文件: {os.path.abspath(OUTPUT_FILENAME_NEGATIVE)}")
    print("="*50)

    # 4. 手动关闭脚本
    input("脚本执行完毕，请按回车键退出...")


if __name__ == "__main__":
    convert_txt_to_excel_dual()
